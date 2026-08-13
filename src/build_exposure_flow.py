"""Flow-based exposure for robustness item 1. See PREREGISTRATION.md s6, s10."""
import pandas as pd
from openpyxl import load_workbook

GFCF_FILE = "data/raw/gfcfbyindustryandasset_2025-11-03.xlsx"
YEARS = [2015, 2016, 2017, 2018, 2019]

QUALIFYING = ["Other_machinery_and_equipment", "ICT"]
OTHER = ["Buildings_and_transfer_costs", "Transport",
         "Intellectual_property_products", "Cultivated_assets"]
DENOMINATOR = QUALIFYING + OTHER

GROUPS = {
    "Solid fuels & oil refining":               ["19"],
    "Chemicals and man made fibres":            ["20", "21"],
    "Metals and metal goods":                   ["24", "25"],
    "Engineering and vehicles":                 ["26", "27", "28", "29", "30"],
    "Food, drink and tobacco":                  ["10 to 12"],
    "Textiles, clothing, leather and footwear": ["13 to 15"],
    "Other manufacturing":                      ["16", "17", "18", "22", "23", "31 to 32", "33"],
    "Agriculture, forestry and fishing":        ["01", "02", "03"],
    "Mining and Quarrying":                     ["05 to 09"],
    "Electricity, gas and water":               ["35", "36", "37 to 39"],
    "Construction":                             ["41 and 43"],
    "Distribution Services":                    ["45", "46", "47"],
    "Transportation and storage":               ["49", "50", "51", "52", "53"],
    "Hotels and restaurants":                   ["55 to 56"],
    "Information and communication":            ["58", "59 to 60", "61", "62 to 63"],
    "Financial intermediation":                 ["64", "65", "66"],
    "Real estate, renting and business":        ["68", "69 to 70", "71", "72", "73",
                                                 "74 to 75", "77", "78", "79", "80 to 82"],
    "Education":                                ["85"],
    "Health and social work":                   ["86", "87 and 88"],
    "Other services":                           ["84", "90 to 92", "93", "94", "95", "96"],
}

wb = load_workbook(GFCF_FILE, read_only=True)

def read_asset(sheet):
    """Return {year: {industry_code: value}} for one asset sheet, current-price table only."""
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    hdrs = [i for i, r in enumerate(rows) if r and r[0] == "SIC07 - INDUSTRY"]
    hdr = hdrs[0]
    end = hdrs[1] if len(hdrs) > 1 else len(rows)
    ncol = sum(1 for c in rows[hdr] if c is not None)
    names = [str(rows[hdr][j]) for j in range(ncol)]
    out = {}
    for r in rows[hdr + 1:end]:
        if isinstance(r[0], int) and r[0] in YEARS:
            out[r[0]] = {names[j]: r[j] for j in range(1, ncol)}
    return out

data = {a: read_asset(a) for a in DENOMINATOR}

n_suppressed = 0
shares = {}
for name, codes in GROUPS.items():
    qual = total = 0.0
    for y in YEARS:
        for c in codes:
            for a in DENOMINATOR:
                v = data[a][y].get(c)
                if not isinstance(v, (int, float)):
                    n_suppressed += 1
                    continue
                total += float(v)
                if a in QUALIFYING:
                    qual += float(v)
    shares[name] = qual / total

result = pd.Series(shares, name="pm_share_flow")
result.index.name = "industry"
result.to_csv("data/processed/exposure_flow.csv")
print(f"suppressed cells: {n_suppressed}")
print(result.sort_values(ascending=False).round(3).to_string())